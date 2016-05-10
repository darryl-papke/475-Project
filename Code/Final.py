__author__ = 'Darryl'

'''
This code is used to classify the gender of a blog author by only using the
words in their blog posts.

This code uses the FIRST CORPUS mentioned in the final write up.

The NLTK and openpyxl libraries are used in this code.

The openpyxl library is used to open and read in the excel files that I used.
'''

from openpyxl import load_workbook
import nltk


'''
The function setup takes in the file name of an Excel sheet and reads
in the entries. The function returns a variable including all posts for
analysis and a variable for both genders that only include post from one
gender.
'''


def setup(filename):
    wb = load_workbook(filename)
    ws = wb.active
    sheet = wb.worksheets[0]
    length = sheet.max_row
    posts = []
    all_posts = ""
    male_posts = ""
    female_posts = ""
    for i in range(1, length + 1):
        posts.append((ws["A" + str(i)].value, ws["B" + str(i)].value,))
    for post in posts:
        all_posts += post[0]
        if post[1] == "M":
            male_posts += post[0]
        else:
            female_posts += post[0]
    return posts, all_posts, male_posts, female_posts

'''
The function blog_features uses a given blog post and develops features
about the author to later either train the classifier or have the classifier
attempt to guess the author's gender.
'''


def blog_features(post):
    features = {}
    tokens = nltk.word_tokenize(post)
    words = [token.lower() for token in tokens if token.isalpha()]
    vocab = set(words)

    for word in top_words:
        features['contains({})'.format(word)] = (word in vocab)

    features['exclamation'] = ("!" in post)
    vocab = len(vocab)
    words = len(words)
    sent = nltk.sent_tokenize(post)
    chars = len([char for sentence in sent for word in sentence for char in word])
    if words == 0:
        features['word_len'] = 1  # post contains one word
        features['vocab_size'] = 1  # post contains one word
        features['sent_len'] = 1  # post contains one word
    else:
        features['word_len'] = round(chars / words)  # average word length
        features['vocab_size'] = round(words / vocab)  # author's vocabulary size
        features['sent_len'] = (True if round(words / len(sent)) == 7 else False)  # average words per sentence
        features['sent_len'] = round(words / len(sent))

    return features


data, total, male, female = setup('dataset.xlsx')
print("data")

'''
Create a male and female list of the most common words used by
both genders.
'''
male = nltk.word_tokenize(male)
top_male = nltk.FreqDist(w.lower() for w in male if w.isalpha())
top_male = top_male.most_common()
top_male = [w for (w, count) in top_male]
print("male")
female = nltk.word_tokenize(female)
top_female = nltk.FreqDist(w.lower() for w in female if w.isalpha())
top_female = top_female.most_common()
top_female = [w for (w, counts) in top_female]
print("female")

'''
Find the top 4000 words that female authors used that don't occur
in male author's blog post and add them to a list (top_words).
The top words list will not include common words like, the, a, and,
will, etc.
'''
top_words = []
for post in top_female:
    if post not in top_male:
        top_words.append(post)
    if len(top_words) == 4000:
        break
top_words = top_words[0:4000]

'''
Find the top 4000 words that male authors used that don't occur
in female author's blog post and add them to the list top_words.
'''
for post in top_male:
    if post not in top_female:
        top_words.append(post)
    if len(top_words) == 8000:
        break
top_words = top_words[0:8000]

print("top_words", len(top_words))

'''
Develop feature sets using the blog_features function and the gender
of the blog's author.
'''
feature_sets = [(blog_features(post[0]), post[1].upper()) for post in data]
size = int(len(feature_sets) * 0.2)
train_set, test_set = feature_sets[size:], feature_sets[:size]

'''
Train, test and print results of the Naive Bayes Classifier using
feature sets.
'''
classifier = nltk.NaiveBayesClassifier.train(train_set)
print(nltk.classify.accuracy(classifier, test_set))
print(classifier.show_most_informative_features(20))

'''
Train, test and print results of the Decision Tree Classifier using
feature sets.

The Decision Tree Classifier was not really used during my research.
'''
# classifier = nltk.DecisionTreeClassifier.train(train_set)
# print(nltk.classify.accuracy(classifier, test_set))
